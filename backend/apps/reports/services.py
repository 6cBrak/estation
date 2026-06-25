"""
Services de rapports : toutes les fonctions retournent des dictionnaires sérialisables.
Aucun modèle propre — agrégation des apps existantes.
"""
from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal

from django.db.models import Count, DecimalField, F, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def _decimal(val) -> Decimal:
    return val if val is not None else Decimal("0")


# ─────────────────────────────────────────────
# Dashboard Super Admin — vue réseau
# ─────────────────────────────────────────────

def network_dashboard(target_date: date | None = None) -> dict:
    """Vue réseau pour le Super Admin : agrégats de toutes les stations."""
    from apps.charges.models import Charge
    from apps.fuel.models import Tank
    from apps.journal.models import StationJournal
    from apps.stations.models import Station

    if target_date is None:
        target_date = timezone.localdate()

    stations = Station.objects.filter(is_active=True).order_by("name")

    journals_map = {
        j.station_id: j
        for j in StationJournal.objects.filter(
            journal_date=target_date, is_active=True
        ).select_related("station")
        .prefetch_related(
            "fuel_lines__nozzle__tank__fuel_type",
            "payment_summary",
        )
    }

    # Charges du jour groupées par station (toutes stations en une seule requête)
    charges_map: dict = {}
    for c in Charge.objects.filter(
        charge_date=target_date, is_active=True, station__is_active=True
    ).select_related("station").order_by("category", "created_at"):
        charges_map.setdefault(c.station_id, []).append(c)

    net_fuel_xof = Decimal("0")
    net_fuel_liters = Decimal("0")
    net_encaisse = Decimal("0")
    net_charges_xof = Decimal("0")
    station_rows = []

    for st in stations:
        j = journals_map.get(st.id)
        st_fuel_xof = Decimal("0")
        st_fuel_liters = Decimal("0")
        st_encaisse = Decimal("0")
        st_status = None
        st_journal_number = None
        pump_rows = []

        if j:
            st_status = j.status
            st_journal_number = j.journal_number
            for line in j.fuel_lines.all():
                if line.sold_volume is not None:
                    st_fuel_liters += line.sold_volume
                if line.amount_xof is not None:
                    st_fuel_xof += line.amount_xof
                pump_rows.append({
                    "label": line.nozzle.label,
                    "fuel_type": line.nozzle.tank.fuel_type.name,
                    "sold_volume": line.sold_volume,
                    "amount_xof": line.amount_xof,
                })

            payment = getattr(j, "payment_summary", None)
            if payment:
                st_encaisse = (
                    payment.cash_amount_xof
                    + payment.tickets_amount_xof
                    + payment.tpe_amount_xof
                    + payment.mobile_money_amount_xof
                    + payment.credit_amount_xof
                )

        # Niveaux des cuves
        tanks = Tank.objects.filter(station=st, is_active=True).select_related("fuel_type").order_by("label")
        tank_rows = []
        for tank in tanks:
            pct = round(float(tank.current_level_liters / tank.capacity_liters * 100), 1) if tank.capacity_liters > 0 else 0
            pct = min(pct, 100)
            tank_rows.append({
                "label": tank.label,
                "fuel_type": tank.fuel_type.name,
                "current_level": tank.current_level_liters,
                "capacity": tank.capacity_liters,
                "pct": pct,
                "is_low": tank.is_low,
            })

        # Dépenses du jour pour cette station
        st_charges = charges_map.get(st.id, [])
        st_charges_xof = sum(
            (c.amount_xof for c in st_charges if c.status == "validated"),
            Decimal("0"),
        )
        charge_rows = [
            {
                "label": c.label,
                "category_display": c.get_category_display(),
                "amount_xof": c.amount_xof,
                "status": c.status,
                "status_display": c.get_status_display(),
            }
            for c in st_charges
        ]

        net_fuel_xof += st_fuel_xof
        net_fuel_liters += st_fuel_liters
        net_encaisse += st_encaisse
        net_charges_xof += st_charges_xof

        station_rows.append({
            "station": st.name,
            "station_code": st.code,
            "station_id": str(st.id),
            "status": st_status,
            "status_display": j.get_status_display() if j else None,
            "journal_number": st_journal_number,
            "fuel_xof": st_fuel_xof,
            "fuel_liters": st_fuel_liters,
            "encaisse_xof": st_encaisse,
            "charges_xof": st_charges_xof,
            "charge_rows": charge_rows,
            "has_charges": len(charge_rows) > 0,
            "pump_rows": pump_rows,
            "tank_rows": tank_rows,
            "has_journal": j is not None,
        })

    return {
        "date": str(target_date),
        "stations_total": len(station_rows),
        "stations_with_journal_today": sum(1 for r in station_rows if r["has_journal"]),
        "net_fuel_liters": net_fuel_liters,
        "net_fuel_xof": net_fuel_xof,
        "net_encaisse_xof": net_encaisse,
        "net_charges_xof": net_charges_xof,
        "stations": station_rows,
    }


# ─────────────────────────────────────────────
# Dashboard Gérant — vue station
# ─────────────────────────────────────────────

def station_dashboard(station, target_date: date | None = None) -> dict:
    """Tableau de bord pour un gérant : données de sa station du jour."""
    from apps.fuel.models import Tank
    from apps.journal.models import JournalSalesRecap, StationJournal
    from apps.sales.models import CashSession

    if target_date is None:
        target_date = timezone.localdate()

    journal = StationJournal.objects.filter(
        station=station, journal_date=target_date, is_active=True
    ).first()

    # Stocks carburant actuels
    tanks = Tank.objects.filter(station=station, is_active=True).select_related("fuel_type")
    tank_stocks = []
    alerts = []
    for t in tanks:
        pct = round(float(t.current_level_liters / t.capacity_liters * 100), 1) if t.capacity_liters > 0 else 0
        tank_stocks.append({
            "label": t.label,
            "fuel_type": t.fuel_type.name,
            "current_level": t.current_level_liters,
            "capacity": t.capacity_liters,
            "pct": pct,
            "is_low": t.is_low,
        })
        if t.is_low:
            alerts.append({
                "type": "tank_low",
                "label": f"Cuve {t.label} ({t.fuel_type.name}) niveau bas — {int(t.current_level_liters)} L",
                "severity": "critical",
            })

    # Journal du jour
    journal_data = None
    if journal:
        fuel_lines = journal.fuel_lines.select_related("nozzle__tank__fuel_type")
        fuel_summary = []
        total_fuel_xof = Decimal("0")
        total_liters = Decimal("0")
        for line in fuel_lines:
            amt = line.amount_xof or Decimal("0")
            vol = line.sold_volume or Decimal("0")
            total_fuel_xof += amt
            total_liters += vol
            fuel_summary.append({
                "pump": line.nozzle.label,
                "fuel_type": line.nozzle.tank.fuel_type.name,
                "sold_volume": line.sold_volume,
                "amount_xof": amt,
                "gauge_diff": line.gauge_diff,
            })
            # Alerte écart jaugeage
            if (
                line.gauge_diff is not None
                and line.theoretical_stock
                and line.theoretical_stock > 0
            ):
                ecart_pct = abs(line.gauge_diff / line.theoretical_stock * 100)
                if ecart_pct > station.gauge_tolerance_pct:
                    alerts.append({
                        "type": "gauge_diff",
                        "label": (
                            f"Écart jaugeage {line.nozzle.label} : "
                            f"{line.gauge_diff:+.1f} L ({ecart_pct:.1f}%)"
                        ),
                        "severity": "warning",
                    })

        recaps = list(journal.sales_recaps.all())
        total_sales_xof = sum(r.daily_value_xof for r in recaps) or Decimal("0")

        payment = getattr(journal, "payment_summary", None)

        # Écart de caisse
        cash_variance = None
        cash_alert = False
        if payment and total_sales_xof > 0:
            cash_variance = payment.total_xof - total_sales_xof
            cash_alert = abs(cash_variance) > station.cash_tolerance_xof
            if cash_alert:
                alerts.append({
                    "type": "cash_variance",
                    "label": f"Écart de caisse : {cash_variance:+.0f} FCFA",
                    "severity": "warning",
                })

        journal_data = {
            "journal_number": journal.journal_number,
            "status": journal.status,
            "status_display": journal.get_status_display(),
            "fuel_summary": fuel_summary,
            "total_fuel_xof": total_fuel_xof,
            "total_liters": total_liters,
            "total_sales_xof": total_sales_xof,
            "cash_variance_xof": cash_variance,
            "cash_alert": cash_alert,
            "payment_summary": {
                "cash": payment.cash_amount_xof if payment else 0,
                "tickets": payment.tickets_amount_xof if payment else 0,
                "tpe": payment.tpe_amount_xof if payment else 0,
                "mobile_money": payment.mobile_money_amount_xof if payment else 0,
                "credit": payment.credit_amount_xof if payment else 0,
                "total": payment.total_xof if payment else 0,
            },
        }

    # Cumul mensuel (tous journaux du mois, toutes catégories)
    first_day = target_date.replace(day=1)
    monthly_agg = JournalSalesRecap.objects.filter(
        journal__station=station,
        journal__journal_date__gte=first_day,
        journal__journal_date__lte=target_date,
        journal__is_active=True,
    ).aggregate(total=Sum("daily_value_xof"))
    monthly_xof = _decimal(monthly_agg["total"])

    # Mois précédent
    last_day_prev = first_day - timedelta(days=1)
    first_day_prev = last_day_prev.replace(day=1)
    prev_agg = JournalSalesRecap.objects.filter(
        journal__station=station,
        journal__journal_date__gte=first_day_prev,
        journal__journal_date__lte=last_day_prev,
        journal__is_active=True,
    ).aggregate(total=Sum("daily_value_xof"))
    prev_month_xof = _decimal(prev_agg["total"])

    # Évolution % vs mois précédent
    evolution_pct = None
    if prev_month_xof > 0:
        evolution_pct = round(float((monthly_xof - prev_month_xof) / prev_month_xof * 100), 1)

    # Sessions de caisse ouvertes en ce moment
    open_sessions = CashSession.objects.filter(
        station=station, status="open", is_active=True
    ).count()

    # Dépenses du jour
    from apps.charges.models import Charge
    charges_qs = Charge.objects.filter(
        station=station, charge_date=target_date, is_active=True
    ).order_by("category", "created_at")

    charges_by_cat: dict[str, dict] = {}
    total_charges_validated = Decimal("0")
    total_charges_pending = Decimal("0")
    charge_rows = []

    for c in charges_qs:
        amt = c.amount_xof
        if c.status == "validated":
            total_charges_validated += amt
        elif c.status == "pending":
            total_charges_pending += amt
        cat = c.category
        if cat not in charges_by_cat:
            charges_by_cat[cat] = {
                "label": c.get_category_display(),
                "total_xof": Decimal("0"),
                "count": 0,
            }
        charges_by_cat[cat]["total_xof"] += amt
        charges_by_cat[cat]["count"] += 1
        charge_rows.append({
            "label": c.label,
            "category_display": c.get_category_display(),
            "amount_xof": amt,
            "status": c.status,
            "status_display": c.get_status_display(),
        })

    return {
        "date": str(target_date),
        "station": station.name,
        "station_code": station.code,
        "tank_stocks": tank_stocks,
        "journal": journal_data,
        "monthly_xof": monthly_xof,
        "prev_month_xof": prev_month_xof,
        "evolution_pct": evolution_pct,
        "open_sessions": open_sessions,
        "alerts": alerts,
        "charges": {
            "total_validated_xof": total_charges_validated,
            "total_pending_xof": total_charges_pending,
            "total_xof": total_charges_validated + total_charges_pending,
            "by_category": list(charges_by_cat.values()),
            "rows": charge_rows,
        },
    }


# ─────────────────────────────────────────────
# Rapport ventes par période
# ─────────────────────────────────────────────

def sales_report(station, date_from: date, date_to: date) -> dict:
    """Rapport des ventes carburant et autres sur une période."""
    from apps.journal.models import JournalFuelLine, JournalSalesRecap, StationJournal

    journals = StationJournal.objects.filter(
        station=station,
        journal_date__gte=date_from,
        journal_date__lte=date_to,
        is_active=True,
    )

    # Ventes carburant par jour
    fuel_by_day: dict[str, dict] = {}
    fuel_lines = JournalFuelLine.objects.filter(
        journal__in=journals,
        index_close__isnull=False,
    ).select_related("journal", "nozzle__tank__fuel_type")

    for line in fuel_lines:
        d = str(line.journal.journal_date)
        if d not in fuel_by_day:
            fuel_by_day[d] = {"liters": Decimal("0"), "xof": Decimal("0")}
        if line.sold_volume:
            fuel_by_day[d]["liters"] += line.sold_volume
        if line.amount_xof:
            fuel_by_day[d]["xof"] += line.amount_xof

    # Autres ventes par jour (depuis JournalSalesRecap)
    other_by_day: dict[str, Decimal] = {}
    recaps = JournalSalesRecap.objects.filter(journal__in=journals).select_related("journal")
    recap_by_category: dict[str, dict] = {}
    for recap in recaps:
        d = str(recap.journal.journal_date)
        other_by_day[d] = other_by_day.get(d, Decimal("0")) + recap.daily_value_xof
        cat = recap.category
        if cat not in recap_by_category:
            recap_by_category[cat] = {
                "label": recap.get_category_display(),
                "total_xof": Decimal("0"),
                "total_qty": Decimal("0"),
            }
        recap_by_category[cat]["total_xof"] += recap.daily_value_xof
        recap_by_category[cat]["total_qty"] += recap.qty

    total_fuel_xof = sum(v["xof"] for v in fuel_by_day.values()) or Decimal("0")
    total_other_xof = sum(
        v["total_xof"] for v in recap_by_category.values()
    ) or Decimal("0")
    total_xof = total_fuel_xof + total_other_xof

    # Jours unifiés (carburant + autres)
    all_dates = sorted(set(fuel_by_day) | set(other_by_day))
    days = []
    for d in all_dates:
        f_xof = fuel_by_day.get(d, {}).get("xof", Decimal("0"))
        o_xof = other_by_day.get(d, Decimal("0"))
        days.append({
            "date": d,
            "fuel_xof": f_xof,
            "other_xof": o_xof,
            "total_xof": f_xof + o_xof,
        })

    # by_category comme dict {label: total_xof} pour le frontend
    by_category_dict = {
        v["label"]: v["total_xof"] for v in recap_by_category.values()
    }

    return {
        "station": station.name,
        "date_from": str(date_from),
        "date_to": str(date_to),
        "total_fuel_xof": total_fuel_xof,
        "total_other_xof": total_other_xof,
        "total_xof": total_xof,
        "total_sales_xof": total_xof,  # alias attendu par le frontend
        "days": days,
        "fuel_by_day": [
            {"date": d, "liters": fuel_by_day[d]["liters"], "xof": fuel_by_day[d]["xof"]}
            for d in sorted(fuel_by_day)
        ],
        "by_category": by_category_dict,
    }


# ─────────────────────────────────────────────
# Rapport stocks
# ─────────────────────────────────────────────

def stock_report(station) -> dict:
    """État des stocks actuels : carburant, lubrifiants, boutique, gaz."""
    from apps.fuel.models import Tank
    from apps.gas.models import GasBottleStock
    from apps.lubricants.models import LubricantStock
    from apps.shop.models import ProductStock

    tanks = Tank.objects.filter(station=station, is_active=True).select_related("fuel_type")
    lub_stocks = LubricantStock.objects.filter(
        station=station, is_active=True
    ).select_related("product__brand")
    product_stocks = ProductStock.objects.filter(
        station=station, is_active=True
    ).select_related("product__category")
    gas_stocks = GasBottleStock.objects.filter(
        station=station, is_active=True
    ).select_related("format")

    def _tank_pct(t) -> float:
        if t.capacity_liters > 0:
            return min(round(float(t.current_level_liters / t.capacity_liters * 100), 1), 100)
        return 0

    tank_rows = [
        {
            "label": t.label,
            "fuel_type": t.fuel_type.name,
            "current_level": t.current_level_liters,
            "current": t.current_level_liters,   # alias frontend
            "capacity": t.capacity_liters,
            "pct": _tank_pct(t),
            "is_low": t.is_low,
        }
        for t in tanks
    ]

    return {
        "station": station.name,
        "fuel_tanks": tank_rows,   # ancien nom conservé
        "tanks": tank_rows,        # alias attendu par le frontend
        "lubricants": [
            {
                "product": str(s.product),
                "quantity": s.quantity,
            }
            for s in lub_stocks
        ],
        "products": [
            {
                "product": s.product.name,
                "category": s.product.category.name,
                "quantity": s.quantity,
                "is_low": s.is_low,
            }
            for s in product_stocks
        ],
        "gas": [
            {
                "format": str(s.format),
                "quantity": s.quantity,
            }
            for s in gas_stocks
        ],
    }


# ─────────────────────────────────────────────
# Rapport caisses
# ─────────────────────────────────────────────

def cash_report(station, date_from: date, date_to: date) -> dict:
    """Rapport des clôtures de caisse sur une période."""
    from apps.sales.models import CashSession

    sessions = CashSession.objects.filter(
        station=station,
        opened_at__date__gte=date_from,
        opened_at__date__lte=date_to,
        is_active=True,
    ).select_related("cashier").order_by("opened_at")

    rows = []
    total_variance = Decimal("0")
    for s in sessions:
        variance = s.variance_xof if hasattr(s, "variance_xof") else None
        rows.append({
            "cashier": s.cashier.get_full_name(),
            "opened_at": str(s.opened_at),
            "closed_at": str(s.closed_at) if s.closed_at else None,
            "status": s.status,
            "opening_amount": s.opening_amount_xof,
            "counted_cash": s.counted_cash_xof,
        })

    return {
        "station": station.name,
        "date_from": str(date_from),
        "date_to": str(date_to),
        "sessions": rows,
        "total_sessions": len(rows),
    }


# ─────────────────────────────────────────────
# Rapport rapprochement pompes / caisse / cuves
# ─────────────────────────────────────────────

def reconciliation_report(station, target_date: date) -> dict:
    """
    Rapprochement : pompes (index) vs récap journal vs encaissements.
    Met en évidence les écarts dépassant la tolérance.
    """
    from apps.journal.models import JournalPaymentSummary, StationJournal

    journal = StationJournal.objects.filter(
        station=station, journal_date=target_date, is_active=True
    ).first()

    if not journal:
        return {
            "station": station.name,
            "date": str(target_date),
            "error": "Aucun journal trouvé pour cette date.",
        }

    # Total ventes pompes (depuis index)
    total_fuel_xof = Decimal("0")
    fuel_details = []
    for line in journal.fuel_lines.select_related("nozzle__tank__fuel_type"):
        amt = line.amount_xof or Decimal("0")
        total_fuel_xof += amt
        fuel_details.append({
            "pump": line.nozzle.label,
            "sold_volume": line.sold_volume,
            "amount_xof": amt,
            "gauge_diff": line.gauge_diff,
            "gauge_diff_alert": (
                abs(line.gauge_diff) > (station.gauge_tolerance_pct / 100 * (line.theoretical_stock or 1))
                if line.gauge_diff and line.theoretical_stock
                else False
            ),
        })

    # Total autres ventes (récap)
    total_recap_xof = sum(
        r.daily_value_xof for r in journal.sales_recaps.all()
    ) or Decimal("0")

    # Total encaissements
    payment = getattr(journal, "payment_summary", None)
    total_payments_xof = payment.total_xof if payment else Decimal("0")

    grand_total_sales = total_fuel_xof + total_recap_xof
    cash_variance = total_payments_xof - grand_total_sales
    cash_alert = abs(cash_variance) > station.cash_tolerance_xof

    # Alertes sous forme de liste de strings pour le frontend
    alerts: list[str] = []
    if cash_alert:
        alerts.append(
            f"Écart de caisse : {cash_variance:+.0f} FCFA "
            f"(tolérance ±{station.cash_tolerance_xof} FCFA)"
        )
    for fd in fuel_details:
        if fd.get("gauge_diff_alert") and fd.get("gauge_diff") is not None:
            alerts.append(
                f"Écart jaugeage pompe {fd['pump']} : {fd['gauge_diff']:+.1f} L"
            )

    # Pompes au format attendu par le frontend {label, calc, journal, diff}
    pumps = [
        {
            "label": fd["pump"],
            "calc": fd.get("sold_volume"),      # volume calculé depuis index
            "amount_xof": fd.get("amount_xof"),
            "journal": fd.get("sold_volume"),   # identique ici (pas de double source)
            "diff": fd.get("gauge_diff"),
        }
        for fd in fuel_details
    ]

    # Paiements sous forme de dict {label: montant}
    payments: dict = {}
    if payment:
        payments = {
            "Espèces": payment.cash_amount_xof,
            "Tickets": payment.tickets_amount_xof,
            "TPE / Carte": payment.tpe_amount_xof,
            "Mobile Money": payment.mobile_money_amount_xof,
            "Crédit": payment.credit_amount_xof,
        }

    return {
        "station": station.name,
        "date": str(target_date),
        "journal_number": journal.journal_number,
        "journal_status": journal.status,
        "fuel_details": fuel_details,
        "total_fuel_xof": total_fuel_xof,
        "total_recap_xof": total_recap_xof,
        "grand_total_sales_xof": grand_total_sales,
        "total_payments_xof": total_payments_xof,
        "cash_variance_xof": cash_variance,
        "cash_alert": cash_alert,
        "cash_tolerance_xof": station.cash_tolerance_xof,
        # Champs attendus par le frontend
        "alerts": alerts,
        "pumps": pumps,
        "payments": payments,
    }


# ─────────────────────────────────────────────
# Export Excel
# ─────────────────────────────────────────────

def export_sales_excel(station, date_from: date, date_to: date) -> bytes:
    """Exporte le rapport ventes en fichier Excel (bytes)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    report = sales_report(station, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2C5F8A")

    # En-tête
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Rapport Ventes — {station.name} — {date_from} au {date_to}"
    ws["A1"].font = Font(bold=True, size=13)

    ws.append([])
    ws.append(["Date", "Litres vendus", "Montant carburant (FCFA)", ""])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    for row in report["fuel_by_day"]:
        ws.append([row["date"], float(row["liters"]), float(row["xof"]), ""])

    ws.append([])
    ws.append(["Catégorie", "Quantité", "Total (FCFA)", ""])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    for cat in report["by_category"]:
        ws.append([cat["label"], float(cat["total_qty"]), float(cat["total_xof"]), ""])

    ws.append([])
    ws.append(["TOTAL CARBURANT", "", float(report["total_fuel_xof"]), ""])
    ws.append(["TOTAL AUTRES", "", float(report["total_other_xof"]), ""])
    ws.append(["TOTAL GÉNÉRAL", "", float(report["total_xof"]), ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
